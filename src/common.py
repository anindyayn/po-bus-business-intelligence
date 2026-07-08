from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

try:
    from sklearn.cluster import KMeans
    from sklearn.ensemble import IsolationForest
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics import accuracy_score, silhouette_score
    from sklearn.model_selection import train_test_split
    from sklearn.preprocessing import StandardScaler
    from sklearn.tree import DecisionTreeClassifier

    SKLEARN_AVAILABLE = True
except Exception:
    KMeans = None
    IsolationForest = None
    TfidfVectorizer = None
    accuracy_score = None
    silhouette_score = None
    train_test_split = None
    StandardScaler = None
    DecisionTreeClassifier = None
    SKLEARN_AVAILABLE = False

try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing

    STATSMODELS_AVAILABLE = True
except Exception:
    ExponentialSmoothing = None
    STATSMODELS_AVAILABLE = False


@dataclass
class Metric:
    info_code: str
    method: str
    metric_name: str
    metric_value: Any
    notes: str


def clean_for_export(df: pd.DataFrame) -> pd.DataFrame:
    clean = df.copy()
    for col in clean.columns:
        if pd.api.types.is_datetime64_any_dtype(clean[col]):
            clean[col] = clean[col].dt.strftime("%Y-%m-%d")
    return clean.replace([np.inf, -np.inf], np.nan)


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            values = [str(cell.value) for cell in column_cells[:80] if cell.value is not None]
            width = min(max([len(v) for v in values] + [10]) + 2, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


def rolling_forecast(series: pd.Series, horizon: int) -> np.ndarray:
    values = series.dropna().tail(14)
    if values.empty:
        return np.zeros(horizon)
    base = values.rolling(window=min(7, len(values)), min_periods=1).mean().iloc[-1]
    return np.repeat(float(base), horizon)


def mean_absolute_percentage_error(actual: pd.Series, predicted: pd.Series) -> float:
    actual_arr = np.asarray(actual, dtype=float)
    pred_arr = np.asarray(predicted, dtype=float)
    mask = actual_arr != 0
    if not mask.any():
        return np.nan
    return float(np.mean(np.abs((actual_arr[mask] - pred_arr[mask]) / actual_arr[mask])))


def safe_ceil(value: float) -> int:
    if pd.isna(value):
        return 0
    return int(math.ceil(max(value, 0)))


def top_value(values: pd.Series) -> Any:
    mode = values.dropna().astype(str).mode()
    if mode.empty:
        return np.nan
    return mode.iloc[0]


def latest_by_date(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    return df.sort_values("date").groupby(group_col, as_index=False).tail(1).copy()


def normalize_0_100(series: pd.Series) -> np.ndarray:
    values = pd.to_numeric(series, errors="coerce").fillna(0).astype(float)
    min_value = values.min()
    max_value = values.max()
    if max_value == min_value:
        return np.zeros(len(values), dtype=float)
    return ((values - min_value) / (max_value - min_value) * 100).to_numpy()


def z_score(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").fillna(0).astype(float)
    std = values.std(ddof=0)
    if std == 0 or pd.isna(std):
        return pd.Series(np.zeros(len(values)), index=series.index)
    return (values - values.mean()) / std


def top_keywords(texts: pd.Series, top_n: int = 6) -> str:
    words: dict[str, int] = {}
    stopwords = {
        "dan",
        "yang",
        "tidak",
        "ada",
        "dengan",
        "untuk",
        "tanpa",
        "keluhan",
        "perjalanan",
    }
    for text in texts.fillna("").astype(str):
        for word in text.lower().split():
            word = "".join(char for char in word if char.isalnum())
            if len(word) < 4 or word in stopwords:
                continue
            words[word] = words.get(word, 0) + 1
    ranked = sorted(words.items(), key=lambda item: item[1], reverse=True)
    return ", ".join(word for word, _ in ranked[:top_n])
